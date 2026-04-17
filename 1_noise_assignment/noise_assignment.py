"""
noise_assignment_final.py
────────────────────────────────────────────────────────────────
功能：
  遍历数据集每个样本，根据其类别查找兼容噪声集合，
  使用配额上限法随机分配一种噪声，保证 25 种噪声
  最终污染的样本数量尽量相等。

使用方式：
  1. 修改 CONFIG 区域的路径和参数
  2. python noise_assignment_final.py

依赖：
  pip install openpyxl pandas numpy
────────────────────────────────────────────────────────────────
"""

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from collections import defaultdict


# ══════════════════════════════════════════════════════════════
# CONFIG：根据实际情况修改这里
# ══════════════════════════════════════════════════════════════

XLSX_PATH   = "noise_compatibility_KS50_VGGSound.xlsx"  # 兼容性表路径
SHEET_NAME  = "KS-50 (50 classes)"                      # 用哪个数据集的 sheet
                                                         # KS-50  → "KS-50 (50 classes)"
                                                         # VGGSound → "VGGSound (309 classes)"
INPUT_CSV   = None          # 你的样本 CSV 路径，None 则使用下方 mock 数据
OUTPUT_CSV  = "dataset_with_noise.csv"   # 输出结果路径
SEED        = 42            # 随机种子，保证结果可复现


# ══════════════════════════════════════════════════════════════
# 25 种噪声的完整列表（顺序与 Excel 一致）
# ══════════════════════════════════════════════════════════════

ALL_NOISES = [
    "V_gaussian_noise", "V_shot_noise", "V_impulse_noise", "V_defocus_blur",
    "V_glass_blur", "V_motion_blur", "V_zoom_blur", "V_snow", "V_frost", "V_fog",
    "V_brightness", "V_contrast", "V_elastic_transform", "V_pixelate", "V_jpeg_compression",
    "A_gaussian_noise", "A_traffic", "A_crowd", "A_rain", "A_thunder", "A_wind",
    "VA_gaussian", "VA_rain",
    "Missing_audio", "Missing_video",
]


# ══════════════════════════════════════════════════════════════
# STEP 1：读取 Excel，构建兼容性字典
# ══════════════════════════════════════════════════════════════

def load_compat_table(xlsx_path: str, sheet_name: str) -> dict:
    """
    返回：{ "pumping_fist": ["V_gaussian_noise", "V_snow", ...], ... }
    """
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    noise_names = list(rows[1][5:])      # 第 2 行第 6 列起是 25 个噪声名
    compat_table = {}
    for row in rows[2:]:                 # 第 3 行起是数据
        class_name = row[1]
        compatible = [noise_names[i] for i, v in enumerate(row[5:]) if v == "✓"]
        compat_table[class_name] = compatible

    return compat_table


# ══════════════════════════════════════════════════════════════
# STEP 2：配额上限法分配噪声
# ══════════════════════════════════════════════════════════════

def assign_noises(samples_df: pd.DataFrame, compat_table: dict, seed: int = 42) -> pd.DataFrame:
    """
    为每个样本分配一种噪声，保证每种噪声分配数量尽量相等。

    算法（配额上限法）：
      1. 计算每种噪声的名额上限 = 总样本数 ÷ 25
      2. 打乱样本顺序（消除同类别样本集中带来的偏差）
      3. 遍历每个样本：
           - 找出该类别兼容的噪声中，还没达到上限的那些
           - 在这些候选里纯随机选一种
           - 该噪声计数 +1
      4. 恢复原始顺序后返回

    samples_df 必须包含列：
      - sample_id  ：样本唯一标识
      - class_name ：类别名（需与 Excel 里的 Class Name 完全一致）
    """
    np.random.seed(seed)

    total = len(samples_df)
    quota = total // 25                  # 每种噪声的名额上限

    noise_counts = defaultdict(int)      # 记录每种噪声已分配数量

    # ── 关键：打乱顺序 ────────────────────────────────────────
    # 同类别样本通常连续存放，不打乱会导致开局分配不均。
    df = samples_df.sample(frac=1, random_state=seed).reset_index(drop=True)

    assignments = []
    fallback_count = 0                   # 触发兜底逻辑的次数（正常应为 0）

    for _, row in df.iterrows():
        class_name = row["class_name"]
        compatible = compat_table.get(class_name)

        if compatible is None:
            raise ValueError(
                f"类别 '{class_name}' 不在兼容性表中。\n"
                f"请确认类别名与 Excel 的 'Class Name' 列完全一致。"
            )

        # 筛选出还有名额的噪声
        available = [n for n in compatible if noise_counts[n] < quota]

        # 兜底：极少数情况下该类别所有兼容噪声都满了（物理约束导致），
        # 此时放宽限制，从完整兼容集里随机选。
        if not available:
            available = compatible
            fallback_count += 1

        chosen = np.random.choice(available)
        noise_counts[chosen] += 1
        assignments.append(chosen)

    df["assigned_noise"] = assignments

    if fallback_count > 0:
        print(f"[提示] 有 {fallback_count} 个样本触发了兜底逻辑（兼容集内所有噪声均已满额），"
              f"这是物理约束导致的正常现象。")

    # 恢复原始顺序
    return df.sort_values("sample_id").reset_index(drop=True)


# ══════════════════════════════════════════════════════════════
# STEP 3：均衡性报告
# ══════════════════════════════════════════════════════════════

def print_balance_report(result_df: pd.DataFrame, compat_table: dict):
    """
    打印均衡性报告。
    对于"物理上限不足"的噪声（兼容样本总数 < 配额），单独标注，
    不算入均衡性判断（属于物理约束，非算法问题）。
    """
    counts = result_df["assigned_noise"].value_counts().reindex(ALL_NOISES, fill_value=0)
    total  = counts.sum()
    quota  = total // 25

    # 计算每种噪声在数据集里最多能分配到多少个样本（物理上限）
    noise_physical_max = defaultdict(int)
    for noises in compat_table.values():
        for n in noises:
            noise_physical_max[n] += 1   # 该噪声兼容的样本类别数
    # 实际物理上限 = 该类别数 × 每类平均样本数（简单估计）
    samples_per_class = total / len(compat_table)
    physical_ceiling  = {n: int(noise_physical_max[n] * samples_per_class)
                         for n in ALL_NOISES}

    # 区分"物理受限"vs"算法问题"
    physically_limited = {n for n in ALL_NOISES if physical_ceiling[n] < quota}

    # CV 只算不受物理约束的噪声
    normal_counts = counts[[n for n in ALL_NOISES if n not in physically_limited]]
    cv = normal_counts.std() / normal_counts.mean() if len(normal_counts) > 0 else 0

    print("\n" + "═" * 72)
    print(f"  均衡性报告  |  总样本={total}  |  配额上限={quota}/种  |  CV={cv:.3f}（排除物理受限项）")
    print("═" * 72)
    print(f"  {'噪声类型':<32} {'分配数':>6}  {'物理上限':>8}  可视化")
    print("─" * 72)

    max_cnt = counts.max()
    for noise, cnt in counts.items():
        bar      = "█" * int(cnt / max_cnt * 22) if max_cnt > 0 else ""
        ceiling  = physical_ceiling[noise]
        if noise in physically_limited:
            flag = f"  ← 物理上限仅 {ceiling}，正常"
        elif cnt < quota * 0.85:
            flag = "  ⚠ 算法分配偏少"
        elif cnt > quota * 1.15:
            flag = "  ⚠ 算法分配偏多"
        else:
            flag = ""
        print(f"  {noise:<32} {cnt:>6}  {ceiling:>8}  {bar}{flag}")

    print("─" * 72)
    print(f"  配额={quota}  最大={counts.max()}  最小={counts.min()}\n")
    print("  结论：", end="")
    if cv < 0.15:
        print("✅ 算法分配均衡，可以使用。")
        if physically_limited:
            names = ", ".join(physically_limited)
            print(f"  ℹ  [{names}] 因兼容类别少，数量偏低属物理约束，非算法问题。")
    else:
        print("⚠  算法分配均衡性较差，请检查兼容性表或样本类别分布。")
    print()


# ══════════════════════════════════════════════════════════════
# 主程序
# ══════════════════════════════════════════════════════════════

if __name__ == "__main__":

    # ── 1. 读取兼容性表 ───────────────────────────────────────
    print(f"[1/4] 读取兼容性表：{XLSX_PATH}  /  sheet: {SHEET_NAME}")
    compat_table = load_compat_table(XLSX_PATH, SHEET_NAME)
    print(f"      共 {len(compat_table)} 个类别")

    # ── 2. 读取 / 构造样本表 ──────────────────────────────────
    print(f"[2/4] 加载样本数据")
    if INPUT_CSV is not None:
        # ── 真实使用：从 CSV 读取 ──────────────────────────────
        # CSV 至少需要两列：sample_id、class_name
        # 例：
        #   sample_id, class_name, video_path, label
        #   0, pumping_fist, /data/v0.mp4, 0
        #   1, petting_cat,  /data/v1.mp4, 1
        samples = pd.read_csv(INPUT_CSV)
        print(f"      共 {len(samples)} 个样本，来自 {INPUT_CSV}")
    else:
        # ── 演示用：生成模拟数据 ───────────────────────────────
        np.random.seed(0)
        class_list = list(compat_table.keys())
        n = 5000                             # 模拟 5000 个样本
        samples = pd.DataFrame({
            "sample_id": range(n),
            "class_name": np.random.choice(class_list, size=n),
        })
        print(f"      使用模拟数据：{len(samples)} 个样本，{len(class_list)} 个类别")

    # ── 3. 分配噪声 ───────────────────────────────────────────
    print(f"[3/4] 分配噪声（配额上限法，seed={SEED}）")
    result = assign_noises(samples, compat_table, seed=SEED)

    # ── 4. 保存 + 报告 ────────────────────────────────────────
    result.to_csv(OUTPUT_CSV, index=False)
    print(f"[4/4] 结果已保存到 {OUTPUT_CSV}")
    print(f"\n前 10 条结果：")
    print(result[["sample_id", "class_name", "assigned_noise"]].head(10).to_string(index=False))

    print_balance_report(result, compat_table)